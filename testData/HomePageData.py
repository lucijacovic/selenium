import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname": "Lucija", "email": "mail@example.com", "gender": "Female"},
                          {"firstname": "Jens", "email": "mail2@example.com", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):

        book = openpyxl.load_workbook("C:\\Users\\lucij\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        dictionary = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "Testcase2":
                for j in range(2, sheet.max_column + 1):
                    # dictionary["email"] = "mail@example.com"
                    dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dictionary]
